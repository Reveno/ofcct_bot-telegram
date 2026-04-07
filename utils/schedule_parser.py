import asyncio
import json
import re
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

from openpyxl import load_workbook

DAY_COUNT = 5

_LESSON_MARKER: str | None = None


def _course_from_sheet_name(sheet_name: str) -> int | None:
    """Напр. «1 курс», «2 курс» → 1, 2. Інакше None."""
    m = re.match(r"^\s*(\d+)\s*курс", sheet_name.strip(), re.IGNORECASE)
    if m:
        return int(m.group(1))
    return None


def _lesson_row_marker() -> str:
    global _LESSON_MARKER
    if _LESSON_MARKER is None:
        path = Path(__file__).resolve().parent.parent / "locales" / "uk.json"
        with open(path, encoding="utf-8") as f:
            data = json.load(f)
        _LESSON_MARKER = str(
            (data.get("parser") or {}).get("lesson_row_marker") or "para"
        )
    return _LESSON_MARKER


def parse_schedule_xlsx(filepath: str) -> list[dict]:
    """
    Parse college schedule Excel file.

    Returns list of dicts:
    {
        'group':         str,
        'day_of_week':   int,   # 1=Mon … 5=Fri
        'lesson_number': int,   # 0 … 5
        'subject':       str,
        'teacher':       str,
        'room':          str,
        'course':        int | None,  # номер курсу з назви листа («1 курс»)
    }
    """
    wb = load_workbook(filepath, read_only=True, data_only=True)
    results = []

    for sheet_name in wb.sheetnames:
        course = _course_from_sheet_name(sheet_name)
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))

        if len(rows) < 14:
            continue

        group_cols: dict[int, str] = {}
        for col_i, val in enumerate(rows[12]):
            if val and isinstance(val, str) and re.search(r"\d", val):
                group_cols[col_i] = val.strip()

        if not group_cols:
            continue

        day_of_week = 0
        prev_lesson_num = -1
        data = rows[13:]

        idx = 0
        while idx < len(data):
            row_a = data[idx]
            cell_b = _cell(row_a, 1)

            if not (
                cell_b
                and isinstance(cell_b, str)
                and _lesson_row_marker() in cell_b
            ):
                idx += 1
                continue

            m = re.search(r"(\d+)", cell_b)
            lesson_num = int(m.group(1)) if m else 0

            if lesson_num == 0 or (prev_lesson_num >= 0 and lesson_num < prev_lesson_num):
                day_of_week += 1

            if day_of_week == 0:
                day_of_week = 1

            if day_of_week > DAY_COUNT:
                break

            prev_lesson_num = lesson_num

            row_b = data[idx + 1] if idx + 1 < len(data) else ()

            for col_i, group_name in group_cols.items():
                subject = _cell(row_a, col_i)
                room = _cell(row_a, col_i + 1)
                teacher = _cell(row_b, col_i)

                if subject:
                    row_dict: dict = {
                        "group": group_name,
                        "day_of_week": day_of_week,
                        "lesson_number": lesson_num,
                        "subject": subject,
                        "teacher": teacher or "",
                        "room": str(room) if room else "",
                    }
                    if course is not None:
                        row_dict["course"] = course
                    results.append(row_dict)

            idx += 2

    wb.close()
    return results


def _cell(row: tuple, idx: int) -> str | None:
    try:
        val = row[idx]
    except (IndexError, TypeError):
        return None
    if val is None:
        return None
    s = str(val).strip()
    return s if s else None


async def parse_schedule_async(filepath: str) -> list[dict]:
    loop = asyncio.get_event_loop()
    with ThreadPoolExecutor() as pool:
        return await loop.run_in_executor(pool, parse_schedule_xlsx, filepath)
