from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

DAY_ALIASES: dict[str, int] = {
    "1": 1,
    "понеділок": 1,
    "пн": 1,
    "2": 2,
    "вівторок": 2,
    "вт": 2,
    "3": 3,
    "середа": 3,
    "ср": 3,
    "4": 4,
    "четвер": 4,
    "чт": 4,
    "5": 5,
    "п'ятниця": 5,
    "пятниця": 5,
    "пт": 5,
    "6": 6,
    "субота": 6,
    "сб": 6,
}


def _norm_day(value: Any) -> int | None:
    if value is None:
        return None
    s = str(value).strip().lower()
    return DAY_ALIASES.get(s)


def parse_consultations_xlsx(path: str) -> list[dict[str, Any]]:
    wb = load_workbook(filename=path, data_only=True)
    ws = wb.active
    rows: list[dict[str, Any]] = []
    # Очікувані колонки: ЦК | Викладач | День | Час | Кабінет | Предмет | Примітки
    for r in ws.iter_rows(min_row=2, values_only=True):
        commission = str(r[0] or "").strip()
        teacher = str(r[1] or "").strip()
        day = _norm_day(r[2])
        time_s = str(r[3] or "").strip()
        room = str(r[4] or "").strip()
        subject = str(r[5] or "").strip()
        notes = str(r[6] or "").strip() if len(r) > 6 else ""
        if not (teacher and day and time_s and room):
            continue
        rows.append(
            {
                "commission": commission,
                "teacher": teacher,
                "day_of_week": int(day),
                "time": time_s,
                "room": room,
                "subject": subject,
                "notes": notes,
            }
        )
    return rows


def parse_consultations_docx(path: str) -> list[dict[str, Any]]:
    try:
        from docx import Document  # type: ignore
    except Exception as e:  # pragma: no cover
        raise RuntimeError("python-docx is required for .docx import") from e

    doc = Document(path)
    out: list[dict[str, Any]] = []
    for table in doc.tables:
        current_commission = ""
        for row_idx, row in enumerate(table.rows):
            cells = [c.text.strip() for c in row.cells]
            if row_idx == 0:
                continue
            if len(cells) < 5:
                continue
            # Формат Word-таблиці: [ЦК/викладач, Пн, Вт, Ср, Чт, Пт, Кабінет]
            # Рядок заголовку ЦК часто дублює назву у всіх клітинках.
            first = (cells[0] if len(cells) > 0 else "").strip()
            if first:
                uniq = {c for c in cells if c}
                if len(uniq) == 1 and len(cells) >= 6:
                    current_commission = first
                    continue
            teacher = first
            room = (cells[6] if len(cells) > 6 else "").strip()
            if not teacher:
                continue
            has_week_grid = len(cells) >= 7 and any(cells[i].strip() for i in range(1, 6))
            if has_week_grid:
                for col_idx in range(1, 6):
                    time_s = cells[col_idx].strip()
                    if not time_s:
                        continue
                    out.append(
                        {
                            "commission": current_commission,
                            "teacher": teacher,
                            "day_of_week": col_idx,  # Пн..Пт
                            "time": time_s,
                            "room": room,
                            "subject": "",
                            "notes": "",
                        }
                    )
                continue

            # Резерв: "xlsx-подібний" рядок у docx.
            commission = teacher
            teacher2 = cells[1] if len(cells) > 1 else ""
            day = _norm_day(cells[2] if len(cells) > 2 else "")
            time_s = cells[3] if len(cells) > 3 else ""
            room2 = cells[4] if len(cells) > 4 else ""
            subject = cells[5] if len(cells) > 5 else ""
            notes = cells[6] if len(cells) > 6 else ""
            if teacher2 and day and time_s and room2:
                out.append(
                    {
                        "commission": commission,
                        "teacher": teacher2,
                        "day_of_week": int(day),
                        "time": time_s,
                        "room": room2,
                        "subject": subject,
                        "notes": notes,
                    }
                )
    if not out:
        raise RuntimeError(
            f"Не вдалося знайти таблиці з консультаціями у {Path(path).name}"
        )
    return out

