"""Генерація порожнього .xlsx для масового імпорту консультацій."""

from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def build_consultations_template_xlsx() -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Консультації"
    headers = ["ЦК", "Викладач", "День", "Час", "Кабінет", "Предмет", "Примітки"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.append(
        [
            "Циклова комісія з ІТ",
            "Іваненко Іван Іванович",
            "Пн",
            "14:00",
            "301",
            "Програмування",
            "Онлайн за домовленістю",
        ]
    )
    ws.append(
        [
            "Циклова комісія з мови",
            "Петренко О.С.",
            "3",
            "15:30",
            "212",
            "",
            "",
        ]
    )
    widths = [28, 26, 10, 10, 12, 24, 28]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws2 = wb.create_sheet("Інструкція")
    instructions = [
        "Заповнюйте дані з другого рядка. Рядок заголовків не змінюйте й не видаляйте.",
        "",
        "День тижня: число 1–6 (1 = понеділок) або текст: Пн, вівторок, ср тощо.",
        "Час: наприклад 14:00 або 14:30.",
        "Обовʼязково: Викладач, День, Час, Кабінет. ЦК, Предмет і Примітки — за потреби.",
        "",
        "Збережіть файл і надішліть у бот: Консультації → Імпорт консультацій.",
        "",
        "Альтернатива: один файл .docx з табличним графіком консультацій (як стандартний Word-шаблон коледжу).",
    ]
    for r, line in enumerate(instructions, start=1):
        ws2.cell(row=r, column=1, value=line)
    ws2.column_dimensions["A"].width = 88

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
